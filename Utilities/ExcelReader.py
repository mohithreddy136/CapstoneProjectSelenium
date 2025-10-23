import openpyxl


# --- MODIFIED --- Added optional row_limit parameter
def get_data_from_excel(file_path, sheet_name, row_limit=None):
    """
    Reads test data from an Excel file.

    :param file_path: Path to the .xlsx file.
    :param sheet_name: Name of the sheet to read from.
    :param row_limit: (Optional) The maximum number of data rows to read (excluding the header).
    :return: A list of lists, where each inner list represents a row of data.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = []

        # Determine the last row to read
        # --- MODIFIED --- Calculate max_row based on row_limit if provided
        max_row_to_read = sheet.max_row
        if row_limit is not None and row_limit > 0:
            # Add 1 because header is skipped (row 2 is the first data row)
            max_row_to_read = min(sheet.max_row, row_limit + 1)

        # Start from the second row (index 2) to skip the header
        # --- MODIFIED --- Use max_row_to_read in the loop
        for row in range(2, max_row_to_read + 1):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                row_data.append(sheet.cell(row=row, column=col).value)
            data.append(row_data)

        # --- NEW --- Print how many rows were read for clarity
        print(f"Read {len(data)} rows of data from {file_path}, sheet '{sheet_name}'.")
        return data

    except FileNotFoundError:
        print(f"Error: The file at {file_path} was not found.")
        return None
    except KeyError:
        print(f"Error: The sheet named '{sheet_name}' was not found in the workbook.")
        return None
    except Exception as e:
        print(f"An unexpected error occurred while reading the Excel file: {e}")
        return None
